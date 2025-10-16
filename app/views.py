import requests
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from .models import Schedule, Appeal, Incident, TrafficJam, Patrol, Camera, NewsCategory, Article, ChatThread, ChatMessage, AccidentType, AccidentData, AccidentStatistics, WeatherData, TrafficForecast, RoadWorks, Detector, VehiclePass, RouteCluster
from django import forms
from django.utils import timezone
import json
from django.db.models import IntegerField, CharField
from django.db.models.functions import Cast
from django.db import models
from django.shortcuts import render, get_object_or_404
import json
from django.utils.dateparse import parse_datetime
from django.views.decorators.http import require_http_methods
from django.db import transaction
from datetime import datetime, timedelta
from collections import defaultdict, Counter

# Simple in-memory cache for OSRM routing between coordinate pairs
_ROUTE_CACHE = {}
_ROUTE_CACHE_MAX = 5000

def _route_cache_set(key, value):
    try:
        if len(_ROUTE_CACHE) > _ROUTE_CACHE_MAX:
            # Drop arbitrary items to keep memory bounded
            for _ in range(1000):
                try:
                    _ROUTE_CACHE.pop(next(iter(_ROUTE_CACHE)))
                except Exception:
                    break
        _ROUTE_CACHE[key] = value
    except Exception:
        pass

def _osrm_route(a_lat, a_lon, b_lat, b_lon):
    """Call OSRM public server to get road-constrained path between two points.
    Returns list of [lat, lon]. Falls back to direct segment on failure.
    """
    key = f"{a_lat:.6f},{a_lon:.6f}|{b_lat:.6f},{b_lon:.6f}"
    if key in _ROUTE_CACHE:
        return _ROUTE_CACHE[key]
    url = f"https://router.project-osrm.org/route/v1/driving/{a_lon},{a_lat};{b_lon},{b_lat}?overview=full&geometries=geojson"
    try:
        resp = requests.get(url, timeout=6)
        if resp.status_code == 200:
            data = resp.json()
            routes = (data or {}).get('routes') or []
            if routes:
                geom = routes[0].get('geometry') or {}
                coords = geom.get('coordinates') or []
                if coords:
                    path = [[lat, lon] for lon, lat in coords]
                    _route_cache_set(key, path)
                    return path
    except Exception:
        pass
    # Fallback straight line
    path = [[a_lat, a_lon], [b_lat, b_lon]]
    _route_cache_set(key, path)
    return path

def main(request):
    categories = NewsCategory.objects.all()
    
    # Получаем данные для верхней панели
    weather = WeatherData.objects.first()
    traffic_forecast = TrafficForecast.objects.first()
    road_works = RoadWorks.objects.first()
    
    # Получаем статистику ДТП
    accident_stats = AccidentStatistics.objects.first()
    
    # Получаем данные для графика ДТП
    accident_types = AccidentType.objects.all()
    accident_data = {}
    for accident_type in accident_types:
        data = list(AccidentData.objects.filter(accident_type=accident_type, year=2024).order_by('month').values_list('count', flat=True))
        # Дополняем данные нулями до 12 месяцев, если данных недостаточно
        while len(data) < 12:
            data.append(0)
        accident_data[accident_type.name] = {
            'color': accident_type.color,
            'data': data
        }
    
    context = {
        'categories': categories,
        'weather': weather,
        'traffic_forecast': traffic_forecast,
        'road_works': road_works,
        'accident_stats': accident_stats,
        'accident_data': accident_data,
    }
    return render(request, 'MainHTML/Main.html', context)


class AppealForm(forms.ModelForm):
    class Meta:
        model = Appeal
        fields = ['name', 'email', 'message']


@csrf_exempt
def appeal_submit(request):
    if request.method == 'POST':
        form = AppealForm(request.POST)
        if form.is_valid():
            form.save()
            # AJAX/JSON response for nicer UX
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.headers.get('accept', '').find('application/json') != -1:
                return JsonResponse({'ok': True})
            return render(request, 'MainHTML/AppealThanks.html')
        # invalid
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.headers.get('accept', '').find('application/json') != -1:
            return JsonResponse({'ok': False, 'errors': form.errors}, status=400)
        return render(request, 'MainHTML/AppealForm.html', {'form': form})

    form = AppealForm()
    return render(request, 'MainHTML/AppealForm.html', {'form': form})


def incidents_list(request):
    incidents = Incident.objects.all().order_by('-occurred_at')
    return render(request, 'MainHTML/IncidentsList.html', {'incidents': incidents})


def incidents_monitoring(request):
    total = Incident.objects.count()
    by_status = (
        Incident.objects
        .values('status')
        .order_by('status')
        .annotate(count=models.Count('id'))
    )
    by_severity = (
        Incident.objects
        .values('severity')
        .order_by('severity')
        .annotate(count=models.Count('id'))
    )
    last_7_days = Incident.objects.filter(occurred_at__gte=timezone.now()-timezone.timedelta(days=7)).count()
    context = {
        'total': total,
        'by_status': list(by_status),
        'by_severity': list(by_severity),
        'last_7_days': last_7_days,
    }
    return render(request, 'MainHTML/IncidentsMonitoring.html', context)


def incidents_api(request):
    qs = Incident.objects.all().order_by('-occurred_at')
    data = []
    for i in qs:
        data.append({
            'id': i.id,
            'title': i.title,
            'description': i.description,
            'occurred_at': i.occurred_at.isoformat(),
            'status': i.status,
            'severity': i.severity,
            'coordinates': i.coordinates or [],
        })
    return JsonResponse({'incidents': data})


def traffic_api(request):
    qs = TrafficJam.objects.all().order_by('-occurred_at')
    data = []
    for t in qs:
        coords = t.coordinates or []
        if len(coords) >= 2:
            data.append({
                'id': t.id,
                'title': t.title,
                'description': t.description,
                'coordinates': coords,
                'severity': t.severity,
                'occurred_at': t.occurred_at.isoformat(),
            })
    return JsonResponse({'traffic': data})


def patrols_api(request):
    qs = Patrol.objects.all().order_by('-created_at')
    data = []
    for p in qs:
        coords = p.coordinates or []
        if len(coords) >= 2:
            data.append({
                'id': p.id,
                'title': p.title,
                'description': p.description,
                'coordinates': coords,
                'radius_m': p.radius_m,
            })
    return JsonResponse({'patrols': data})


def cameras_api(request):
    qs = Camera.objects.all().order_by('-created_at')
    data = []
    for c in qs:
        coords = c.coordinates or []
        if isinstance(coords, str):
            try:
                coords = json.loads(coords)
            except Exception:
                coords = []
        if len(coords) >= 2:
            data.append({
                'id': c.id,
                'name': c.name,
                'description': c.description,
                'coordinates': coords,
            })
    return JsonResponse({'cameras': data})


def detectors_api(request):
    qs = Detector.objects.all().order_by('id')
    data = []
    for d in qs:
        coords = d.coordinates or []
        if isinstance(coords, str):
            try:
                coords = json.loads(coords)
            except Exception:
                coords = []
        data.append({
            'id': d.id,
            'external_id': d.external_id,
            'name': d.name,
            'coordinates': coords,
        })
    return JsonResponse({'detectors': data})


def news_categories(request):
    categories = NewsCategory.objects.all()
    return render(request, 'MainHTML/NewsCategories.html', {'categories': categories})


def news_category_articles(request, category_id):
    category = get_object_or_404(NewsCategory, id=category_id)
    articles = Article.objects.filter(category=category).order_by('-created_at')
    return render(request, 'MainHTML/NewsArticles.html', {'category': category, 'articles': articles})


def news_article_detail(request, article_id):
    article = get_object_or_404(Article, id=article_id)
    return render(request, 'MainHTML/NewsArticleDetail.html', {'article': article})


@csrf_exempt
def chat_send_message(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = request.POST

    content = (payload or {}).get('content', '').strip()
    subject = (payload or {}).get('subject', '').strip()
    sender_type = (payload or {}).get('sender', 'user')
    
    if not content:
        return JsonResponse({'ok': False, 'error': 'empty content'}, status=400)

    # Use or create thread for authenticated user; guests get separate thread per session
    thread = None
    if request.user.is_authenticated:
        thread_id = (payload or {}).get('thread_id')
        if thread_id:
            thread = ChatThread.objects.filter(id=thread_id, is_closed=False).first()
        if not thread:
            thread = ChatThread.objects.create(user=request.user, subject=subject)
    else:
        # store thread id in session
        thread_id = request.session.get('guest_thread_id')
        if thread_id:
            thread = ChatThread.objects.filter(id=thread_id, user__isnull=True, is_closed=False).first()
        if not thread:
            thread = ChatThread.objects.create(user=None, subject=subject)
            request.session['guest_thread_id'] = thread.id

    # Определяем отправителя
    if sender_type == 'admin':
        sender = ChatMessage.Sender.ADMIN
    else:
        sender = ChatMessage.Sender.USER

    msg = ChatMessage.objects.create(thread=thread, sender=sender, content=content)
    return JsonResponse({'ok': True, 'thread_id': thread.id, 'message_id': msg.id})


def chat_thread_messages(request):
    thread_id = request.GET.get('thread_id')
    if request.user.is_authenticated:
        thread = ChatThread.objects.filter(id=thread_id, user=request.user).first() if thread_id else ChatThread.objects.filter(user=request.user).order_by('-created_at').first()
    else:
        guest_thread_id = request.session.get('guest_thread_id')
        if not guest_thread_id:
            return JsonResponse({'messages': [], 'thread_id': None})
        thread = ChatThread.objects.filter(id=guest_thread_id, user__isnull=True).first()

    if not thread:
        return JsonResponse({'messages': [], 'thread_id': None})

    messages = list(thread.messages.order_by('created_at').values('id', 'sender', 'content', 'created_at'))
    return JsonResponse({'thread_id': thread.id, 'messages': messages, 'is_authenticated': request.user.is_authenticated})


# API для управления данными верхней панели
@csrf_exempt
def update_weather(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            weather, created = WeatherData.objects.get_or_create(pk=1)
            weather.temperature = data.get('temperature', 0)
            weather.description = data.get('description', '')
            weather.icon = data.get('icon', 'fa-sun')
            weather.save()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)


@csrf_exempt
def update_traffic_forecast(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            forecast, created = TrafficForecast.objects.get_or_create(pk=1)
            forecast.speed = data.get('speed', 0)
            forecast.save()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)


@csrf_exempt
def update_road_works(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            works, created = RoadWorks.objects.get_or_create(pk=1)
            works.count = data.get('count', 0)
            works.save()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)


# API для управления статистикой ДТП
@csrf_exempt
def update_accident_stats(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            stats, created = AccidentStatistics.objects.get_or_create(pk=1)
            stats.total_accidents = data.get('total_accidents', 0)
            stats.injured = data.get('injured', 0)
            stats.killed = data.get('killed', 0)
            stats.year = data.get('year', 2024)
            stats.save()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)


@csrf_exempt
def update_accident_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            accident_type_id = data.get('accident_type_id')
            month = data.get('month')
            count = data.get('count', 0)
            year = data.get('year', 2024)
            
            accident_data, created = AccidentData.objects.get_or_create(
                accident_type_id=accident_type_id,
                month=month,
                year=year,
                defaults={'count': count}
            )
            if not created:
                accident_data.count = count
                accident_data.save()
            
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)


def get_accident_types(request):
    types = list(AccidentType.objects.values('id', 'name', 'color'))
    return JsonResponse({'types': types})


@csrf_exempt
def accident_types_api(request):
    if request.method == 'GET':
        types = list(AccidentType.objects.values('id', 'name', 'color'))
        return JsonResponse({'types': types})
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            accident_type = AccidentType.objects.create(
                name=data.get('name'),
                color=data.get('color', '#36A2EB')
            )
            return JsonResponse({'ok': True, 'id': accident_type.id})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            accident_type = AccidentType.objects.get(id=data.get('id'))
            accident_type.name = data.get('name', accident_type.name)
            accident_type.color = data.get('color', accident_type.color)
            accident_type.save()
            return JsonResponse({'ok': True})
        except AccidentType.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Accident type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            accident_type = AccidentType.objects.get(id=data.get('id'))
            accident_type.delete()
            return JsonResponse({'ok': True})
        except AccidentType.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Accident type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


def get_accident_data(request):
    year = request.GET.get('year', 2024)
    data = {}
    for accident_type in AccidentType.objects.all():
        data[accident_type.name] = {
            'id': accident_type.id,
            'color': accident_type.color,
            'data': list(AccidentData.objects.filter(
                accident_type=accident_type, 
                year=year
            ).order_by('month').values_list('count', flat=True))
        }
    return JsonResponse(data)


# API для управления новостными категориями
@csrf_exempt
def news_categories_api(request):
    if request.method == 'GET':
        categories = list(NewsCategory.objects.values('id', 'name', 'image_url'))
        return JsonResponse({'categories': categories})
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            category = NewsCategory.objects.create(
                name=data.get('name'),
                image_url=data.get('image_url', '')
            )
            return JsonResponse({'ok': True, 'id': category.id})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            category_id = data.get('id')
            category = NewsCategory.objects.get(id=category_id)
            category.name = data.get('name', category.name)
            category.image_url = data.get('image_url', category.image_url)
            category.save()
            return JsonResponse({'ok': True})
        except NewsCategory.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Category not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            category_id = data.get('id')
            category = NewsCategory.objects.get(id=category_id)
            category.delete()
            return JsonResponse({'ok': True})
        except NewsCategory.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Category not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


# API для управления статьями
@csrf_exempt
def news_articles_api(request):
    if request.method == 'GET':
        articles = list(Article.objects.select_related('category').values(
            'id', 'title', 'author', 'summary', 'content', 'created_at',
            'category__name', 'cover_image_url', 'category_id'
        ))
        return JsonResponse({'articles': articles})
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            category = NewsCategory.objects.get(id=data.get('category_id'))
            article = Article.objects.create(
                category=category,
                title=data.get('title'),
                author=data.get('author', ''),
                summary=data.get('summary', ''),
                content=data.get('content'),
                cover_image_url=data.get('cover_image_url', '')
            )
            return JsonResponse({'ok': True, 'id': article.id})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            article_id = data.get('id')
            article = Article.objects.get(id=article_id)
            article.title = data.get('title', article.title)
            article.author = data.get('author', article.author)
            article.summary = data.get('summary', article.summary)
            article.content = data.get('content', article.content)
            article.cover_image_url = data.get('cover_image_url', article.cover_image_url)
            if data.get('category_id'):
                category = NewsCategory.objects.get(id=data.get('category_id'))
                article.category = category
            article.save()
            return JsonResponse({'ok': True})
        except Article.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Article not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            article_id = data.get('id')
            article = Article.objects.get(id=article_id)
            article.delete()
            return JsonResponse({'ok': True})
        except Article.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Article not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


# API для управления обращениями
@csrf_exempt
def appeals_api(request):
    if request.method == 'GET':
        filter_type = request.GET.get('filter', 'all')
        appeals = Appeal.objects.all()
        
        if filter_type == 'pending':
            appeals = appeals.filter(is_reviewed=False)
        elif filter_type == 'reviewed':
            appeals = appeals.filter(is_reviewed=True)
            
        appeals = list(appeals.values('id', 'name', 'email', 'message', 'is_reviewed', 'created_at'))
        return JsonResponse({'appeals': appeals})
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


@csrf_exempt
def appeal_toggle(request, appeal_id):
    if request.method == 'POST':
        try:
            appeal = Appeal.objects.get(id=appeal_id)
            appeal.is_reviewed = not appeal.is_reviewed
            appeal.save()
            return JsonResponse({'ok': True})
        except Appeal.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Appeal not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


# API для управления чатами
@csrf_exempt
def chat_threads_api(request):
    if request.method == 'GET':
        threads = []
        # Показываем только чаты текущего пользователя или гостевые чаты
        if request.user.is_authenticated:
            # Для авторизованных пользователей показываем только их чаты
            user_threads = ChatThread.objects.filter(user=request.user)
        else:
            # Для гостей показываем только их чаты по session
            guest_thread_id = request.session.get('guest_thread_id')
            if guest_thread_id:
                user_threads = ChatThread.objects.filter(id=guest_thread_id, user__isnull=True)
            else:
                user_threads = ChatThread.objects.none()
        
        for thread in user_threads:
            threads.append({
                'id': thread.id,
                'subject': thread.subject,
                'user_name': thread.user.get_username() if thread.user else 'Гость',
                'message_count': thread.messages.count(),
                'is_closed': thread.is_closed,
                'created_at': thread.created_at.isoformat()
            })
        return JsonResponse({'threads': threads})
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


@csrf_exempt
def chat_thread_toggle(request, thread_id):
    if request.method == 'POST':
        try:
            thread = ChatThread.objects.get(id=thread_id)
            thread.is_closed = not thread.is_closed
            thread.save()
            return JsonResponse({'ok': True})
        except ChatThread.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Thread not found'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)


# API для админов - показывает все чаты
@csrf_exempt
def admin_chat_threads_api(request):
    if request.method == 'GET':
        # Проверяем, что пользователь авторизован (в реальном приложении нужно проверить права админа)
        if not request.user.is_authenticated:
            return JsonResponse({'ok': False, 'error': 'Authentication required'}, status=401)
        
        threads = []
        for thread in ChatThread.objects.all().order_by('-created_at'):
            threads.append({
                'id': thread.id,
                'subject': thread.subject,
                'user_name': thread.user.get_username() if thread.user else 'Гость',
                'message_count': thread.messages.count(),
                'is_closed': thread.is_closed,
                'created_at': thread.created_at.isoformat()
            })
        return JsonResponse({'threads': threads})
    return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)

# API для аутентификации
from django.contrib.auth import authenticate, login, logout


@csrf_exempt
def user_login(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')
            
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return JsonResponse({
                    'success': True,
                    'user': {
                        'id': user.id,
                        'username': user.username,
                        'role': 'admin' if user.is_superuser else 'editor'
                    }
                })
            else:
                return JsonResponse({'success': False, 'error': 'Invalid credentials'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@csrf_exempt
def user_logout(request):
    if request.method == 'POST':
        logout(request)
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@csrf_exempt
def auth_status(request):
    if request.method == 'GET':
        if request.user.is_authenticated:
            return JsonResponse({
                'authenticated': True,
                'user': {
                    'id': request.user.id,
                    'username': request.user.username,
                    'role': 'admin' if request.user.is_superuser else 'editor'
                }
            })
        else:
            return JsonResponse({'authenticated': False})
    
    return JsonResponse({'authenticated': False})


@csrf_exempt
def get_accident_stats(request):
    if request.method == 'GET':
        try:
            accident_stats = AccidentStatistics.objects.first()
            if accident_stats:
                return JsonResponse({
                    'accident_stats': {
                        'total_accidents': accident_stats.total_accidents,
                        'injured': accident_stats.injured,
                        'killed': accident_stats.killed,
                        'year': accident_stats.year
                    }
                })
            else:
                return JsonResponse({
                    'accident_stats': {
                        'total_accidents': 0,
                        'injured': 0,
                        'killed': 0,
                        'year': 2024
                    }
                })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def get_weather_data(request):
    if request.method == 'GET':
        try:
            weather = WeatherData.objects.first()
            if weather:
                return JsonResponse({
                    'weather': {
                        'temperature': weather.temperature,
                        'description': weather.description,
                        'icon': weather.icon
                    }
                })
            else:
                return JsonResponse({
                    'weather': {
                        'temperature': 0,
                        'description': 'Данные не загружены',
                        'icon': 'fa-sun'
                    }
                })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def get_traffic_forecast(request):
    if request.method == 'GET':
        try:
            traffic = TrafficForecast.objects.first()
            if traffic:
                return JsonResponse({
                    'traffic_forecast': {
                        'speed': traffic.speed
                    }
                })
            else:
                return JsonResponse({
                    'traffic_forecast': {
                        'speed': 0
                    }
                })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def get_road_works(request):
    if request.method == 'GET':
        try:
            road_works = RoadWorks.objects.first()
            if road_works:
                return JsonResponse({
                    'road_works': {
                        'count': road_works.count
                    }
                })
            else:
                return JsonResponse({
                    'road_works': {
                        'count': 0
                    }
                })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ---------------- Traffic analytics: ingestion, co-movement, clustering ----------------

@csrf_exempt
@require_http_methods(["POST"])
def ingest_detectors(request):
    """Accept CSV or XLSX with detectors.
    Supported formats:
      - 4 columns: external_id, name, lat, lon
      - 3 columns: name, lat, lon (name will also be used as external_id)
    Creates/updates Detector entries.
    """
    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'ok': False, 'error': 'file is required'}, status=400)

    created, updated = 0, 0
    try:
        # detectors expected order: external_id, name, lat, lon
        if upload.name.lower().endswith('.csv'):
            import io, csv
            text = io.TextIOWrapper(upload.file, encoding='utf-8')
            reader = csv.reader(text)
            rows = []
            first = True
            for cols in reader:
                if first:
                    first = False  # skip header row entirely
                    continue
                rows.append(cols)
        elif upload.name.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            try:
                from openpyxl import load_workbook
            except Exception:
                return JsonResponse({'ok': False, 'error': 'openpyxl not installed'}, status=400)
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        else:
            return JsonResponse({'ok': False, 'error': 'Unsupported file type'}, status=400)

        with transaction.atomic():
            for cols in rows:
                # Normalize row to one of the supported schemas
                cols = list(cols)
                # Trim whitespace from string cells
                cols = [c.strip() if isinstance(c, str) else c for c in cols]

                external_id = ''
                name = ''
                lat = None
                lon = None

                if len(cols) >= 4:
                    # external_id, name, lat, lon
                    external_id = str(cols[0] or '').strip()
                    name = str(cols[1] or '')
                    lat = cols[2]
                    lon = cols[3]
                elif len(cols) >= 3:
                    # name, lat, lon (use name as external_id)
                    name = str(cols[0] or '')
                    external_id = name.strip()
                    lat = cols[1]
                    lon = cols[2]
                else:
                    # Not enough columns
                    continue

                if not external_id:
                    continue

                coords = []
                try:
                    if lat is not None and lon is not None and str(lat) != '' and str(lon) != '':
                        coords = [float(lat), float(lon)]
                except Exception:
                    coords = []

                det, was_created = Detector.objects.update_or_create(
                    external_id=external_id,
                    defaults={'name': name or external_id, 'coordinates': coords}
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        return JsonResponse({'ok': True, 'created': created, 'updated': updated})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def ingest_passes(request):
    """Accept CSV/XLSX with vehicle detections: detector_id, timestamp, vehicle_id, speed(optional)."""
    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'ok': False, 'error': 'file is required'}, status=400)

    created = 0
    detector_cache = {d.external_id: d.id for d in Detector.objects.all()}

    try:
        # passes expected order: detector_id, timestamp, vehicle_id, speed(optional)
        if upload.name.lower().endswith('.csv'):
            import io, csv
            text = io.TextIOWrapper(upload.file, encoding='utf-8')
            reader = csv.reader(text)
            rows = []
            first = True
            for cols in reader:
                if first:
                    first = False  # skip header row entirely
                    continue
                rows.append(cols)
        elif upload.name.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            try:
                from openpyxl import load_workbook
            except Exception:
                return JsonResponse({'ok': False, 'error': 'openpyxl not installed'}, status=400)
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        else:
            return JsonResponse({'ok': False, 'error': 'Unsupported file type'}, status=400)

        with transaction.atomic():
            for cols in rows:
                det_id = cols[0] if len(cols) > 0 else None
                if det_id is None:
                    continue
                det_id = str(det_id)
                detector_pk = detector_cache.get(det_id)
                if not detector_pk:
                    # auto-create detector without coordinates
                    det = Detector.objects.create(external_id=det_id, name='')
                    detector_cache[det.external_id] = det.id
                    detector_pk = det.id

                ts_raw = cols[1] if len(cols) > 1 else None
                if isinstance(ts_raw, datetime):
                    ts = ts_raw
                else:
                    ts = None
                    if ts_raw is not None and ts_raw != '':
                        try:
                            ts = parse_datetime(str(ts_raw)) or datetime.fromisoformat(str(ts_raw))
                        except Exception:
                            ts = None

                vehicle_id = str((cols[2] if len(cols) > 2 else '') or '').strip()
                if not vehicle_id or not ts:
                    continue
                speed = cols[3] if len(cols) > 3 else None
                try:
                    speed = float(speed) if speed is not None and speed != '' else None
                except Exception:
                    speed = None

                VehiclePass.objects.create(
                    detector_id=detector_pk,
                    vehicle_id=vehicle_id,
                    timestamp=ts,
                    speed_kmh=speed,
                )
                created += 1

        return JsonResponse({'ok': True, 'created': created})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


def _vehicle_path(vehicle_id, start=None, end=None):
    qs = VehiclePass.objects.filter(vehicle_id=vehicle_id)
    if start:
        qs = qs.filter(timestamp__gte=start)
    if end:
        qs = qs.filter(timestamp__lte=end)
    qs = qs.order_by('timestamp').values('detector_id', 'timestamp', 'speed_kmh')
    path = [(r['detector_id'], r['timestamp'], r['speed_kmh']) for r in qs]
    return path


@require_http_methods(["GET"])
def vehicle_path_api(request):
    """Get vehicle path for a specific vehicle in a time window.
    Params: vehicle_id, start (ISO), end (ISO).
    Returns: path as list of detector IDs.
    """
    def _parse_iso_to_naive(value: str):
        if not value:
            return None
        try:
            dt = parse_datetime(value)
            if dt is None:
                value2 = value.replace('Z', '+00:00')
                dt = datetime.fromisoformat(value2)
        except Exception:
            try:
                dt = datetime.fromisoformat(value)
            except Exception:
                dt = None
        if dt is None:
            return None
        try:
            from django.utils import timezone as dj_tz
            if dj_tz.is_aware(dt):
                dt = dj_tz.make_naive(dt)
        except Exception:
            pass
        return dt

    vehicle_id = request.GET.get('vehicle_id')
    if not vehicle_id:
        return JsonResponse({'ok': False, 'error': 'vehicle_id is required'}, status=400)
    
    start = request.GET.get('start')
    end = request.GET.get('end')
    start_dt = _parse_iso_to_naive(start) if start else None
    end_dt = _parse_iso_to_naive(end) if end else None
    
    try:
        path_data = _vehicle_path(vehicle_id, start_dt, end_dt)
        # Extract just the detector IDs for the path
        path = [item[0] for item in path_data]
        return JsonResponse({'ok': True, 'path': path})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@require_http_methods(["GET"])
def comovement_api(request):
    """Find vehicles co-moving with target vehicle.
    Params: vehicle_id, k (min consecutive nodes), dt (sec tolerance), max_lead (nodes).
    Optional: start, end ISO datetime.
    """
    vehicle_id = request.GET.get('vehicle_id')
    if not vehicle_id:
        return JsonResponse({'ok': False, 'error': 'vehicle_id is required'}, status=400)
    k = int(request.GET.get('k', 3))
    dt = int(request.GET.get('dt', 300))  # seconds
    max_lead = int(request.GET.get('max_lead', 2))
    start = request.GET.get('start')
    end = request.GET.get('end')
    start_dt = parse_datetime(start) if start else None
    end_dt = parse_datetime(end) if end else None

    target_path = _vehicle_path(vehicle_id, start_dt, end_dt)
    if not target_path:
        return JsonResponse({'ok': True, 'matches': []})

    # Build index: detector_id -> list of (timestamp, vehicle_id)
    det_ids = list({d for d, _, _ in target_path})
    index = defaultdict(list)
    window_start = target_path[0][1] - timedelta(seconds=dt)
    window_end = target_path[-1][1] + timedelta(seconds=dt)
    for vp in VehiclePass.objects.filter(detector_id__in=det_ids, timestamp__gte=window_start, timestamp__lte=window_end).values('detector_id', 'timestamp', 'vehicle_id'):
        if vp['vehicle_id'] == vehicle_id:
            continue
        index[vp['detector_id']].append((vp['timestamp'], vp['vehicle_id']))

    # For each candidate vehicle, walk along target path and check alignment
    candidate_to_matches = defaultdict(list)
    for det_id, ts, _ in target_path:
        for ts2, cand in index.get(det_id, []):
            if abs((ts2 - ts).total_seconds()) <= dt:
                candidate_to_matches[cand].append((det_id, ts, ts2))

    results = []
    for cand, matches in candidate_to_matches.items():
        # Sort by target timestamp and check for at least k consecutive detector matches
        matches.sort(key=lambda x: x[1])
        longest = 1
        current = 1
        lead_balance = 0  # positive if target ahead
        for i in range(1, len(matches)):
            prev_det, prev_t, prev_t2 = matches[i-1]
            det, t, t2 = matches[i]
            if det == prev_det:
                continue
            if (t - prev_t).total_seconds() >= 0:
                current += 1
                # update lead balance
                lead_balance += 1 if (t - t2).total_seconds() < 0 else -1
            else:
                current = 1
                lead_balance = 0
            longest = max(longest, current)
        if longest >= k and abs(lead_balance) <= max_lead:
            start_t = matches[0][1]
            end_t = matches[-1][1]
            results.append({
                'vehicle_id': cand,
                'matched_nodes': longest,
                'start_time': start_t.isoformat(),
                'end_time': end_t.isoformat(),
                'nodes': [m[0] for m in matches],
            })

    results.sort(key=lambda r: r['matched_nodes'], reverse=True)
    return JsonResponse({'ok': True, 'matches': results})


@require_http_methods(["GET"])
def cluster_routes_api(request):
    """Cluster routes in a time window; return top N popular paths with stats.
    Params: start, end (ISO), top (N), min_len (min nodes).
    """
    def _parse_iso_to_naive(value: str):
        if not value:
            return None
        # Accept 'Z' suffix and timezone offsets; convert to naive in server local time
        try:
            dt = parse_datetime(value)
            if dt is None:
                # Fallback: replace 'Z' with '+00:00' for Python fromisoformat
                value2 = value.replace('Z', '+00:00')
                dt = datetime.fromisoformat(value2)
        except Exception:
            # Last resort: try plain fromisoformat without changes
            try:
                dt = datetime.fromisoformat(value)
            except Exception:
                dt = None
        if dt is None:
            return None
        # Make naive
        try:
            from django.utils import timezone as dj_tz
            if dj_tz.is_aware(dt):
                dt = dj_tz.make_naive(dt)
        except Exception:
            pass
        return dt

    start = request.GET.get('start')
    end = request.GET.get('end')
    if not start or not end:
        return JsonResponse({'ok': False, 'error': 'start and end are required'}, status=400)
    start_dt = _parse_iso_to_naive(start)
    end_dt = _parse_iso_to_naive(end)
    if start_dt is None or end_dt is None:
        return JsonResponse({'ok': False, 'error': 'invalid datetime format'}, status=400)
    top_n = int(request.GET.get('top', 10))
    min_len = int(request.GET.get('min_len', 3))

    # Build per-vehicle ordered paths in window
    passes = VehiclePass.objects.filter(timestamp__gte=start_dt, timestamp__lte=end_dt).order_by('vehicle_id', 'timestamp').values('vehicle_id', 'detector_id', 'timestamp', 'speed_kmh')
    vehicle_to_path = defaultdict(list)
    for r in passes:
        vehicle_to_path[r['vehicle_id']].append((r['detector_id'], r['timestamp'], r['speed_kmh']))

    # Normalize paths by removing duplicates in a row and too short
    path_counter = Counter()
    path_stats = {}
    for vid, path in vehicle_to_path.items():
        if not path:
            continue
        normalized = []
        prev = None
        for det, ts, sp in path:
            if prev is None or prev != det:
                normalized.append((det, ts, sp))
            prev = det
        if len(normalized) < min_len:
            continue
        key = tuple([d for d, _, _ in normalized])
        path_counter[key] += 1
        # accumulate stats
        start_ts = normalized[0][1]
        end_ts = normalized[-1][1]
        duration = (end_ts - start_ts).total_seconds()
        speeds = [sp for _, _, sp in normalized if sp is not None]
        if key not in path_stats:
            path_stats[key] = {'vehicles': 0, 'durations': [], 'speeds': []}
        path_stats[key]['vehicles'] += 1
        path_stats[key]['durations'].append(duration)
        path_stats[key]['speeds'].extend(speeds)

    most_common = path_counter.most_common(top_n)
    clusters = []
    hours = max((end_dt - start_dt).total_seconds() / 3600.0, 1e-6)
    for key, count in most_common:
        stats = path_stats.get(key, {'vehicles': 0, 'durations': [], 'speeds': []})
        avg_time = sum(stats['durations']) / len(stats['durations']) if stats['durations'] else None
        avg_speed = sum(stats['speeds']) / len(stats['speeds']) if stats['speeds'] else None
        clusters.append({
            'path': list(key),
            'vehicle_count': count,
            'intensity_per_hour': count / hours,
            'avg_speed_kmh': avg_speed,
            'avg_travel_seconds': avg_time,
        })

    return JsonResponse({'ok': True, 'clusters': clusters})


def traffic_analytics_page(request):
    return render(request, 'MainHTML/TrafficAnalytics.html')


@require_http_methods(["POST"])
def route_snap_api(request):
    """Accept JSON with segments: [[lat, lon], [lat, lon], ...].
    Returns a polyline snapped to roads using OSRM.
    Also supports multiple segments as { segments: [ [[lat,lon],...], [[lat,lon],...] ] }.
    """
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'invalid JSON'}, status=400)

    def _route_for_segment(seg):
        if not isinstance(seg, list) or len(seg) < 2:
            return []
        merged = []
        for i in range(len(seg) - 1):
            a = seg[i]
            b = seg[i+1]
            if not (isinstance(a, list) and isinstance(b, list) and len(a) == 2 and len(b) == 2):
                continue
            a_lat, a_lon = float(a[0]), float(a[1])
            b_lat, b_lon = float(b[0]), float(b[1])
            part = _osrm_route(a_lat, a_lon, b_lat, b_lon)
            if i == 0:
                merged.extend(part)
            else:
                merged.extend(part[1:] if len(part) > 1 else part)
        return merged

    if isinstance(payload, dict) and 'segments' in payload:
        out = []
        for seg in payload.get('segments') or []:
            out.append(_route_for_segment(seg))
        return JsonResponse({'ok': True, 'polylines': out})

    seg = payload if isinstance(payload, list) else []
    polyline = _route_for_segment(seg)
    return JsonResponse({'ok': True, 'polyline': polyline})
